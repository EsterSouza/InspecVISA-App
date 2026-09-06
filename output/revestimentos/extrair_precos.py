"""Lê a planilha oficial da SINAPI e escreve precos.py com um valor por estado.

Fonte: qa/SINAPI-2026-07.zip, baixado do canal de downloads da Caixa. Dentro dele,
SINAPI_Referência_2026_07.xlsx, aba CSD: relatório de custos de composições com
encargos sociais SEM DESONERAÇÃO. Cada UF ocupa uma coluna de custo; os valores
são custos de serviço, sem BDI.

Rode `python extrair_precos.py` só quando trocar o mês de referência.
"""
import io,re,zipfile
from pathlib import Path
import openpyxl

RAIZ=Path(__file__).resolve().parent
PACOTE=RAIZ/'qa'/'SINAPI-2026-07.zip'
# Os códigos que a biblioteca cita. Puxar a planilha inteira encheria a página de
# 10 mil serviços que não têm nada a ver com revestimento de sala de saúde.
# Ficaram de fora, conferido nesta referência: piso laminado (98683), carpete em
# placa e em manta (106782, 106783) e as divisórias removíveis (102238, 102248,
# 102448). As composições existem no catálogo, mas a SINAPI não publicou custo
# para elas em estado nenhum em 07/2026. Não é lacuna do Rio: é lacuna do país.
CODIGOS=['87263','87262','87257','87256','87251','102494','101727','106790',
'87265','104611','88485','88497','88489','96358','96368',
'88496','88484','88488','96113','96114','104757','96116',
'86895','106780','106772','86900',
'88650','98685','101742','98688','102496']
NOMES={'AC':'Acre','AL':'Alagoas','AM':'Amazonas','AP':'Amapá','BA':'Bahia','CE':'Ceará','DF':'Distrito Federal','ES':'Espírito Santo','GO':'Goiás','MA':'Maranhão','MG':'Minas Gerais','MS':'Mato Grosso do Sul','MT':'Mato Grosso','PA':'Pará','PB':'Paraíba','PE':'Pernambuco','PI':'Piauí','PR':'Paraná','RJ':'Rio de Janeiro','RN':'Rio Grande do Norte','RO':'Rondônia','RR':'Roraima','RS':'Rio Grande do Sul','SC':'Santa Catarina','SE':'Sergipe','SP':'São Paulo','TO':'Tocantins'}

def abrir():
 # A coluna do código guarda uma fórmula: lida com data_only ela volta zero. Por
 # isso o arquivo é aberto duas vezes, uma para as fórmulas e outra para os valores.
 with zipfile.ZipFile(PACOTE) as z:
  nome=next(n for n in z.namelist() if 'Refer' in n and n.endswith('.xlsx'))
  bruto=z.read(nome)
 formulas=openpyxl.load_workbook(io.BytesIO(bruto),read_only=True,data_only=False)['CSD']
 valores=openpyxl.load_workbook(io.BytesIO(bruto),read_only=True,data_only=True)['CSD']
 return list(formulas.iter_rows(values_only=True)),list(valores.iter_rows(values_only=True))

def main():
 formulas,linhas=abrir()
 referencia=emissao=''
 colunas={}
 for row in linhas[:10]:
  celulas=['' if c is None else str(c).strip() for c in row]
  texto=' '.join(celulas)
  if 'Mês de Referência' in texto:referencia=celulas[celulas.index('Mês de Referência:')+1]
  if 'Data de emissão' in texto:emissao=celulas[celulas.index('Data de emissão:')+1]
  # A linha das UFs é a que traz as 27 siglas; a coluna da sigla é a do custo.
  achadas={c:i for i,c in enumerate(celulas) if c in NOMES}
  if len(achadas)==27 and not colunas:colunas=achadas
 assert colunas,'não achei a linha de UFs na aba CSD'
 alvo=set(CODIGOS)
 precos={}
 for row,formula in zip(linhas,formulas):
  # O código vem da fórmula da coluna 1, um HYPERLINK que termina em ,<código>).
  # Tem outro número antes, o deslocamento do OFFSET: o que vale é o último.
  numeros=re.findall(r',(\d+)\)',str(formula[1]))
  celulas=[numeros[-1]] if numeros else []
  achado=next((c for c in celulas[:4] if c in alvo),None)
  if not achado or achado in precos:continue
  valores={}
  for uf,col in colunas.items():
   try:v=float(str(row[col]).replace(',','.')) if row[col] not in (None,'') else 0.0
   except ValueError:v=0.0
   if v>0:valores[uf]=round(v,2)
  precos[achado]=valores
 faltando=[c for c in CODIGOS if c not in precos]
 assert not faltando,f'códigos não encontrados na planilha: {faltando}'
 saida=['"""Preços SINAPI por estado. Gerado por extrair_precos.py; não editar à mão."""',
 f'REFERENCIA={referencia!r}',f'EMISSAO={emissao!r}',
 'UFS='+repr(sorted(NOMES.items())),'PRECOS={']
 for c in CODIGOS:
  saida.append(f' {c!r}:'+repr(dict(sorted(precos[c].items())))+',')
 saida.append('}')
 (RAIZ/'precos.py').write_text('\n'.join(saida)+'\n',encoding='utf-8')
 vazios=[c for c in CODIGOS if not precos[c]]
 print(f'{len(CODIGOS)} composições, referência {referencia}, emissão {emissao}')
 print('sem preço em estado nenhum:',vazios or 'nenhuma')
 for c in CODIGOS:
  if len(precos[c])<27:print(f'  {c}: preço em {len(precos[c])} de 27 estados')

if __name__=='__main__':main()
